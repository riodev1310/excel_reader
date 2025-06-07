from openpyxl import load_workbook

wb = load_workbook("company_data.xlsx")
ws = wb["Visitors data"]

column_pricing_id = ws["G"]

if ws['H1'].value != "Pricing Plan":
    ws.insert_cols(8)  # Chèn cột mới tại vị trí cột 8 (giữa "Pricing Plan id" và "City id")
    ws['H1'] = "Pricing Plan"  # Thêm tiêu đề cho cột mới
    
    
for index, cell in enumerate(column_pricing_id, start=1):
    if cell.value == 1: 
        ws[f"H{index}"] = "Light"
    elif cell.value == 2: 
        ws[f"H{index}"] = "Advanced"
    elif cell.value == 3:
        ws[f"H{index}"] = "Premium"
    else: 
        ws[f"H{index}"] = "NONE"

wb.save("company_data_copy.xlsx")