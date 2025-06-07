from openpyxl import load_workbook

def update_pricing_plan():
    try:
        # Mở file Excel
        wb = load_workbook("company_data_init.xlsx")
        
        # Kiểm tra sheet "Visitors data" có tồn tại không
        if "Visitors data" not in wb.sheetnames:
            raise ValueError("Sheet 'Visitors data' không tồn tại trong file!")
        
        ws = wb["Visitors data"]
        
        # Chèn cột mới nếu chưa có
        if ws['H1'].value != "Pricing Plan":
            ws.insert_cols(8)  # Chèn cột mới tại vị trí cột 8 (giữa "Pricing Plan id" và "City id")
            ws['H1'] = "Pricing Plan"  # Thêm tiêu đề cho cột mới
        
        # Lấy cột Pricing Plan id (cột G)
        column_pricing_id = ws["G"]
        
        # Duyệt qua các hàng, bỏ qua hàng tiêu đề (hàng 1)
        for index, cell in enumerate(column_pricing_id, start=1):
            if index == 1:  # Bỏ qua hàng tiêu đề
                continue
            if cell.value == 1: 
                ws[f"H{index}"] = "Light"
            elif cell.value == 2: 
                ws[f"H{index}"] = "Advanced"
            elif cell.value == 3:
                ws[f"H{index}"] = "Premium"
            else: 
                ws[f"H{index}"] = "NONE"
        
        # Lưu file
        wb.save("company_data_copy.xlsx")
        print("Đã cập nhật file thành công! File được lưu tại 'company_data_copy.xlsx'.")
    
    except FileNotFoundError:
        print("Lỗi: File 'company_data.xlsx' không tồn tại!")
    except Exception as e:
        print(f"Lỗi: {str(e)}")

if __name__ == "__main__":
    update_pricing_plan()