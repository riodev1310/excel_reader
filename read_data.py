from openpyxl import load_workbook 

wb = load_workbook("book_selling_records.xlsx")
ws = wb["Books-Selling-Records-"]

print("Reading data from the excel file") 
for row in ws.iter_rows(values_only=True):
    print(row) 