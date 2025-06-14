from openpyxl import Workbook, load_workbook
import os


excel_file = "visitor_data.xlsx"

if os.path.exists(excel_file):
    workbook = load_workbook(excel_file)
    worksheet = workbook["Sheet1"]
    if worksheet["H1"] != "Pricing Plan":
        worksheet.insert_cols(8)
        worksheet["H1"] = "Pricing Plan"
else: 
    workbook = Workbook() 
    worksheet = workbook.active 
    headers = ["Visitor id", "Channel id", "Device type", "Bought", "Order id", "Product id", 
               "Pricing plan id", "City id", "Day of week (num)", "Day of week"]
    worksheet.append(headers)



# for row in range(2, worksheet.max_row + 1):
#     pricing_plan_id = worksheet[f"G{row}"].value
#     if pricing_plan_id == 1: 
#         worksheet[f"H{row}"] = "Light"
#     elif pricing_plan_id == 2:
#         worksheet[f"H{row}"] = "Advanced"
#     elif pricing_plan_id == 3:
#         worksheet[f"H{row}"] = "Premium"
#     else: 
#         worksheet[f"H{row}"] = "NONE"
    
row_num = 2
for row_values in worksheet.iter_rows(min_row = 2, values_only=True):
    pricing_plan_id = row_values[6]
    if pricing_plan_id == 1: 
        worksheet[f"H{row_num}"] = "Light"
    elif pricing_plan_id == 2:
        worksheet[f"H{row_num}"] = "Advanced"
    elif pricing_plan_id == 3:
        worksheet[f"H{row_num}"] = "Premium"
    else: 
        worksheet[f"H{row_num}"] = "NONE"
    row_num += 1 
workbook.save("visitor_data_pricing_plan.xlsx")