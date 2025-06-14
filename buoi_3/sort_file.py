import openpyxl

# Hàm Bubble Sort ổn định cho mảng 2D dựa trên cột chính và cột phụ
def bubble_sort_2d_stable(arr, primary_column, secondary_column):
    n = len(arr)
    for i in range(n):
        for j in range(0, n - i - 1):
            # So sánh giá trị tại cột chính (Day of week (num))
            if arr[j][primary_column] > arr[j + 1][primary_column]:
                arr[j], arr[j + 1] = arr[j + 1], arr[j]
            # Nếu giá trị cột chính bằng nhau, so sánh cột phụ (Visitor id)
            elif arr[j][primary_column] == arr[j + 1][primary_column]:
                if arr[j][secondary_column] > arr[j + 1][secondary_column]:
                    arr[j], arr[j + 1] = arr[j + 1], arr[j]
    return arr

# Đọc file Excel
workbook = openpyxl.load_workbook('visitor_data.xlsx')
sheet = workbook['Sheet1']

# Lấy dữ liệu từ sheet (bỏ header)
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    data.append(list(row))

# Sắp xếp dữ liệu theo cột 'Day of week (num)' (cột 8, index 7) và 'Visitor id' (cột 1, index 0) làm cột phụ
sorted_data = bubble_sort_2d_stable(data, 8, 0)

# Ghi lại dữ liệu đã sắp xếp vào sheet (giữ header)
for i, row in enumerate(sorted_data, start=2):
    for j, value in enumerate(row):
        sheet.cell(row=i, column=j+1, value=value)

# Lưu file Excel đã sắp xếp
workbook.save('visitor_data_sorted.xlsx')

print("Dữ liệu đã được sắp xếp và lưu vào 'visitor_data_sorted.xlsx'")