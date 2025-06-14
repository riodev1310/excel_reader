from openpyxl import load_workbook

wb = load_workbook("book_selling_records.xlsx")
ws = wb["Books-Selling-Records-"]

column_item_price = ws["W"]

ws["AI1"] = "New Total" 

# for row in range(2, ws.max_row + 1):
#     quantity = ws[f"V{row}"].value
#     price = ws[f"W{row}"].value 
#     shipping_price = ws[f"X{row}"].value
    
#     if quantity is not None and price is not None:
#         total = quantity * price + shipping_price
#     else: 
#         total = 0
    
#     ws[f"AI{row}"] = total 
    
row_num = 2
for row_values in ws.iter_rows(min_row = 2, values_only=True):
    quantity = row_values[21]
    price = row_values[22]
    shipping_price = row_values[23]
    
    if quantity is not None and price is not None:
        total = quantity * price + shipping_price
    else: 
        total = 0
    
    ws[f"AI{row_num}"] = total 
    
    row_num += 1
    # print(quantity, price, shipping_price)

wb.save("book_selling_record.xlsx") 