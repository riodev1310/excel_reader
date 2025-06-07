from openpyxl import load_workbook

wb = load_workbook("book_selling_records.xlsx")
ws = wb["Books-Selling-Records-"]

column_item_price = ws["W"]

for cell in column_item_price:
    print(cell.value) 