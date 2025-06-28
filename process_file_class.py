from openpyxl import Workbook, load_workbook
import re
from docx import Document
import os

class CoverLetterProcessor:
    def __init__(self, folder_path, excel_file):
        """Initialize with folder path and Excel file name."""
        self.folder_path = folder_path
        self.excel_file = excel_file
        self.patterns = {
            "Họ và tên": r"Họ và tên: (.*?)\s+Nam/Nữ",
            "Giới tính": r"Nam/Nữ: (.*?)$",
            "Ngày sinh": r"Sinh ngày: (.*?)\s+Nơi sinh",
            "Nơi sinh": r"Nơi sinh: (.*?)$",
            "Nguyên quán": r"Nguyên quán: (.*?)$",
            "Hộ khẩu": r"Nơi đăng ký hộ khẩu thường trú: (.*?)$",
            "Chỗ ở hiện nay": r"Chỗ ở hiện nay: (.*?)$",
            "Điện thoại": r"Điện thoại liên hệ: (.*?)$",
            "Dân tộc": r"Dân tộc: (.*?)\s+Tôn giáo",
            "CCCD/CMND": r"Số CCCD/CMND: (.*?)\s+Cấp ngày",
            "Ngày cấp": r"Cấp ngày: (.*?)\s+Nơi cấp",
            "Nơi cấp": r"Nơi cấp: (.*?)$",
            "Trình độ văn hóa": r"Trình độ văn hóa: (.*?)$",
            "Sở trường": r"Sở trường: (.*?)$"
        }
        self.headers = [
            "Họ và tên", "Giới tính", "Ngày sinh", "Nơi sinh", "Nguyên quán",
            "Hộ khẩu", "Chỗ ở hiện nay", "Điện thoại", "Dân tộc", "CCCD/CMND",
            "Ngày cấp", "Nơi cấp", "Trình độ văn hóa", "Sở trường"
        ]
        self.wb = None
        self.ws = None

    def initialize_excel(self):
        """Initialize or load the Excel workbook."""
        if os.path.exists(self.excel_file):
            self.wb = load_workbook(self.excel_file)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "User Information"
            self.ws.append(self.headers)

    def read_docx(self, file_path):
        """Read text from a .docx file."""
        try:
            doc = Document(file_path)
            return "\n".join(para.text for para in doc.paragraphs)
        except Exception as e:
            print(f"Error reading {file_path}: {e}")
            return None

    def extract_info(self, text):
        """Extract information from text using regex patterns."""
        info = {}
        for key, pattern in self.patterns.items():
            match = re.search(pattern, text, re.MULTILINE)
            if match:
                info[key] = match.group(1).strip()
        return info

    def process_documents(self):
        """Process all .docx files in the folder and save to Excel."""
        self.initialize_excel()
        doc_files = os.listdir(self.folder_path)

        for file_name in doc_files:
            file_path = os.path.join(self.folder_path, file_name)
            print(f"Processing {file_name}...")

            if not file_name.endswith(".docx"):
                print(f"Unsupported file format for {file_name}. Skipping...")
                continue

            document_text = self.read_docx(file_path)
            if document_text is None:
                print(f"Failed to read {file_name}. Skipping...")
                continue

            data = self.extract_info(document_text)
            values = [data.get(header, '') for header in self.headers]
            self.ws.append(values)

        self.wb.save(self.excel_file)
        print("Process run successfully!!!")

def main():
    folder_path = "./cover_letters"  # Replace with your folder path
    excel_file = "so_yeu_ly_lich_test.xlsx"
    processor = CoverLetterProcessor(folder_path, excel_file)
    processor.process_documents()

if __name__ == "__main__":
    main()