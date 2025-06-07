from openpyxl import Workbook, load_workbook
import re
from docx import Document  # For .docx files
import os

# Function to extract text from .docx file
def read_docx(file_path):
    try:
        doc = Document(file_path)
        full_text = []
        for para in doc.paragraphs:
            full_text.append(para.text)
        return "\n".join(full_text)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return None


# Function to extract information using regex
def extract_info(text):
    info = {}
    patterns = {
        "Họ và tên": r"Họ và tên: (.*?)\s+Nam/Nữ",
        "Giới tính": r"Nam/Nữ: (.*?)$",
        "Ngày sinh": r"Sinh ngày: (.*?)\s+Nơi sinh",
        "Nơi sinh": r"Nơi sinh: (.*?)$",
        "Nguyên quán": r"Nguyên quán: (.*?)$",
        "Hộ khẩu": r"Nơi đăng ký hộ khẩu thường trú: (.*?)$",
        "Chỗ ở hiện nay": r"Chỗ ở hiện nay: (.*?)$",
        "Điện thoại": r"Điện thoại liên hệ: (.*?)$",
        "Dân tộc": r"Dân tộc: (.*?)\s+Tôn giáo",
        "CCCD/CMND": r"Số CCCD/CMND: (.*?)\s+Cấp ngày",
        "Ngày cấp": r"Cấp ngày: (.*?)\s+Nơi cấp",
        "Nơi cấp": r"Nơi cấp: (.*?)$",
        "Trình độ văn hóa": r"Trình độ văn hóa: (.*?)$",
        "Sở trường": r"Sở trường: (.*?)$"
    }
    
    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.MULTILINE)
        if match:
            info[key] = match.group(1).strip()
    
    return info

# Folder containing the .doc or .docx files
folder_path = "./cover_letters"  # Replace with your folder path

# Define excel file to save information
excel_file = "so_yeu_ly_lich.xlsx"

# Find all .doc and .docx files in the folder
doc_files = os.listdir(folder_path)


if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "User Information"
    # Defines headers
    headers = [
        "Họ và tên", "Giới tính", "Ngày sinh", "Nơi sinh", "Nguyên quán",
        "Hộ khẩu", "Chỗ ở hiện nay", "Điện thoại", "Dân tộc", "CCCD/CMND",
        "Ngày cấp", "Nơi cấp", "Trình độ văn hóa", "Sở trường"
    ]
    ws.append(headers)


# Process each document and write to Excel
for file_path in doc_files:
    print(f"Processing {file_path}...")
    
    # Read the document
    if file_path.endswith(".docx"):
        document_text = read_docx(f"./cover_letters/{file_path}")
    else:
        print(f"Unsupported file format for {file_path}. Skipping...")
        continue
    
    if document_text is None:
        print(f"Failed to read {file_path}. Skipping...")
        continue
    
    # Extract information
    data = extract_info(document_text)
    
    # Extract values from data
    values = []
    for value in data.values():
        values.append(value)
    
    # Append to the excel file
    ws.append(values)

# Save Excel file
wb.save(excel_file)
print("Process run successfully!!!")