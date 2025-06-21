import openpyxl
import json
from openpyxl.utils.exceptions import InvalidFileException

# Load the workbook
try:
    wb = openpyxl.load_workbook('company data test.xlsx')
except FileNotFoundError:
    print("Error: company data.xlsx not found.")
    exit(1)
except InvalidFileException:
    print("Error: Invalid Excel file format.")
    exit(1)

# Verify worksheet exists
if "Visitors data" not in wb.sheetnames:
    print("Error: Worksheet 'Visitors data' not found in the workbook.")
    exit(1)
ws = wb["Visitors data"]

# Define dictionaries for mapping
channel_map = {
    1: 'Google SEO',
    2: 'Google Ads',
    3: 'Bing',
    4: 'Facebook',
    5: 'Instagram',
    6: 'Twitter',
    10: 'Unknown',
    12: 'Unknown'
}

product_map = {
    1: 'Pintury',
    2: 'Pixo',
    3: 'Floati',
    4: 'Artista',
    5: 'SprayPix',
    -1: 'NONE'
}

pricing_plan_map = {
    1: 'Light',
    2: 'Advanced',
    3: 'Premium',
    -1: 'NONE'
}

city_map = {
    1: 'Ho Chi Minh City',
    2: 'Hanoi',
    3: 'Da Nang',
    4: 'Huế',
    5: 'Đông Hà',
    6: 'Đà Lạt',
    7: 'Hội An',
    8: 'Buôn Ma Thuột',
    9: 'Haiphong',
    10: 'Nha Trang'
}

# Load alphabet positions from JSON
try:
    with open("alphabet.json", 'r', encoding='utf-8') as file:
        alphabet_positions = json.load(file)
except FileNotFoundError:
    print("Error: alphabet.json not found.")
    exit(1)
except json.JSONDecodeError:
    print("Error: Invalid JSON format in alphabet.json.")
    exit(1)

def insert_new_column(new_column_name, insert_after):
    """Insert a new column after a specific column if it doesn't already exist."""
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == insert_after:
            col_idx = col + 1  # Insert right after
            if ws.cell(row=1, column=col_idx).value != new_column_name:
                ws.insert_cols(col_idx)
                ws.cell(row=1, column=col_idx).value = new_column_name
            return True
    print(f"Error: Column {insert_after} not found.")
    return False

# Define new columns and their insertion positions
new_columns = [
    {'name': 'Channel', 'insert_after': 'Channel id'},
    {'name': 'Products', 'insert_after': 'Product id'},
    {'name': 'Pricing Plan', 'insert_after': 'Pricing plan id'},
    {'name': 'City', 'insert_after': 'City id'}
]

# Insert new columns
for col_info in new_columns:
    if not insert_new_column(
        new_column_name=col_info['name'],
        insert_after=col_info['insert_after']
    ):
        print(f"Skipping data mapping due to missing {col_info['insert_after']} column.")
        exit(1)

# Find the column indices for all columns after insertion
header_row = 1
col_map = {}
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=header_row, column=col).value
    if header:  # Only add non-empty headers
        col_map[header] = col

# Verify all required columns exist
required_columns = ['Channel id', 'Product id', 'Pricing plan id', 'City id', 'Channel', 'Products', 'Pricing Plan', 'City']
for col in required_columns:
    if col not in col_map:
        print(f"Error: Required column {col} not found after column insertion.")
        exit(1)

# Map the data for each row
for row in range(2, ws.max_row + 1):
    # Get ID values
    channel_id = ws.cell(row=row, column=col_map['Channel id']).value
    product_id = ws.cell(row=row, column=col_map['Product id']).value
    pricing_plan_id = ws.cell(row=row, column=col_map['Pricing plan id']).value
    city_id = ws.cell(row=row, column=col_map['City id']).value

    # Map IDs to values, with default for invalid IDs
    ws.cell(row=row, column=col_map['Channel']).value = channel_map.get(channel_id, 'Unknown')
    ws.cell(row=row, column=col_map['Products']).value = product_map.get(product_id, 'Unknown')
    ws.cell(row=row, column=col_map['Pricing Plan']).value = pricing_plan_map.get(pricing_plan_id, 'Unknown')
    ws.cell(row=row, column=col_map['City']).value = city_map.get(city_id, 'Unknown')

# Save the modified workbook
try:
    # wb.save('company_data_mapped.xlsx')
    wb.save("company data test.xlsx")
    print("Data mapping completed and saved to company_data_mapped.xlsx")
except Exception as e:
    print(f"Error saving workbook: {e}")
    exit(1)